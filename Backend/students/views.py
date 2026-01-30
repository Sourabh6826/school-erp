from rest_framework import viewsets, filters, status
from rest_framework.parsers import MultiPartParser
import openpyxl
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Sum, Q, Max
from .models import Student
from .serializers import StudentSerializer
from fees.models import FeeHead, FeeTransaction, FeeAmount, GlobalFeeSetting, StudentFeeEnrollment
from datetime import datetime

class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.all()
    serializer_class = StudentSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'student_id']
    filterset_fields = ['student_class']

    @action(detail=False, methods=['get'])
    def stats(self, request):
        try:
            session = request.query_params.get('session')
            student_class = request.query_params.get('student_class')
            installment = request.query_params.get('installment')
            date_from = request.query_params.get('date_from')
            date_to = request.query_params.get('date_to', datetime.now().strftime('%Y-%m-%d'))

            # 1. Fetch Students
            students_qs = Student.objects.all()
            if student_class:
                students_qs = students_qs.filter(student_class=student_class)
            
            # Use values to speed up base stats
            student_list = list(students_qs.values('id', 'student_class', 'status', 'has_transport', 'transport_fee_head_id'))
            
            total_students = len(student_list)
            active_count = sum(1 for s in student_list if s['status'] == 'Active')
            tc_count = sum(1 for s in student_list if s['status'] == 'TC')

            # 2. Fetch Global Settings
            global_settings = GlobalFeeSetting.objects.filter(session=session).first()
            g_inst_count = global_settings.installment_count if global_settings else 1
            if g_inst_count == 0: g_inst_count = 1

            # 3. Fetch All Fee Heads for the session
            heads = list(FeeHead.objects.filter(session=session))
            
            # 4. Bulk Fetch Fee Amounts
            amounts_qs = FeeAmount.objects.filter(fee_head__in=heads)
            amounts_map = {(a.fee_head_id, a.class_name): float(a.amount) for a in amounts_qs}

            # 5. Bulk Fetch Transactions (Filtered by date/installment if needed)
            transactions_qs = FeeTransaction.objects.filter(fee_head__in=heads)
            if student_class:
                transactions_qs = transactions_qs.filter(student__student_class=student_class)
            if date_from:
                transactions_qs = transactions_qs.filter(payment_date__gte=date_from)
            if date_to:
                transactions_qs = transactions_qs.filter(payment_date__lte=date_to)
            if installment:
                transactions_qs = transactions_qs.filter(installment_number=installment)

            total_collected = float(transactions_qs.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0)

            # 6. Bulk Fetch Enrollments (Opt-Outs) - Only needed if we want perfect accuracy for "Expected"
            # For stats, we might simplify or be precise. Let's be precise.
            enrollments_qs = StudentFeeEnrollment.objects.filter(session=session)
            # Map: (student_id, fee_head_id, installment_number) -> is_enrolled
            enroll_map = {(e.student_id, e.fee_head_id, e.installment_number): e.is_enrolled for e in enrollments_qs}

            # 7. Calculate "Expected" in memory
            total_expected = 0
            for s in student_list:
                # Get heads applicable to this class
                for h in heads:
                    amt = amounts_map.get((h.id, s['student_class']))
                    if amt is None: continue
                    
                    # Transport logic
                    if h.is_transport_fee and (not s['has_transport'] or s['transport_fee_head_id'] != h.id):
                        continue
                    
                    inst_count = 1 if h.frequency == 'ONCE' else g_inst_count
                    
                    # If specific installment requested
                    if installment:
                        target_inst = int(installment)
                        if target_inst <= inst_count:
                            if enroll_map.get((s['id'], h.id, target_inst), True):
                                total_expected += (amt / inst_count)
                    else:
                        # Sum all valid installments
                        for i in range(1, inst_count + 1):
                            if enroll_map.get((s['id'], h.id, i), True):
                                total_expected += (amt / inst_count)
            
            total_pending = float(total_expected) - total_collected

            return Response({
                'total_students': total_students,
                'active_students': active_count,
                'tc_students': tc_count,
                'total_collected': total_collected,
                'total_pending': max(0, total_pending), # Prevent negative due to overpayments
            })
        except Exception as e:
            import traceback
            print(f"Error in stats: {str(e)}")
            print(traceback.format_exc())
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['get'])
    def pending_fees(self, request):
        try:
            session = request.query_params.get('session')
            student_class = request.query_params.get('student_class')
            show_all = request.query_params.get('show_all') == 'true'
            student_id_param = request.query_params.get('student_id')

            # 1. Fetch Students
            students = Student.objects.all()
            if student_class:
                students = students.filter(student_class=student_class)
            if student_id_param:
                students = students.filter(id=student_id_param)
            
            # 2. Fetch Global Settings
            global_settings = GlobalFeeSetting.objects.filter(session=session).first()
            g_inst_count = global_settings.installment_count if global_settings else 1
            if g_inst_count == 0: g_inst_count = 1

            # 3. Fetch All Fee Heads for the session
            heads_qs = FeeHead.objects.filter(session=session)
            heads = list(heads_qs) # Force evaluation once
            head_ids = [h.id for h in heads]
            
            # 4. Bulk Fetch Fee Amounts
            amounts_qs = FeeAmount.objects.filter(fee_head__in=heads)
            # Map: (fee_head_id, class_name) -> amount
            amounts_map = {(a.fee_head_id, a.class_name): float(a.amount) for a in amounts_qs}

            # 5. Bulk Fetch Transactions
            transactions_qs = FeeTransaction.objects.filter(student__in=students, fee_head__in=heads)
            # Map: (student_id, fee_head_id, installment_number) -> total_paid
            trans_map = {}
            for t in transactions_qs:
                key = (t.student_id, t.fee_head_id, t.installment_number)
                trans_map[key] = trans_map.get(key, 0) + float(t.amount_paid)

            # 6. Bulk Fetch Enrollments (Opt-Outs)
            enrollments_qs = StudentFeeEnrollment.objects.filter(student__in=students, fee_head__in=heads, session=session)
            # Map: (student_id, fee_head_id, installment_number) -> is_enrolled
            enroll_map = {(e.student_id, e.fee_head_id, e.installment_number): e.is_enrolled for e in enrollments_qs}

            fee_detail_list = []
            
            # Convert students to list to prevent re-querying count etc
            student_list = list(students)

            for student in student_list:
                # Pre-filter heads applicable to this class in memory
                student_class_heads = [h for h in heads if (h.id, student.student_class) in amounts_map]
                
                total_expected = 0
                total_paid = 0
                
                installment_data = {i: {'heads': {}} for i in range(1, g_inst_count + 1)}

                for head in student_class_heads:
                    try:
                        # Transport logic
                        if head.is_transport_fee and (not student.has_transport or student.transport_fee_head_id != head.id):
                            continue

                        total_amt = amounts_map.get((head.id, student.student_class), 0)
                        if total_amt == 0: continue

                        inst_count = 1 if head.frequency == 'ONCE' else g_inst_count
                        if inst_count <= 0: inst_count = 1
                        
                        inst_amt = total_amt / inst_count
                        
                        display_name = "Transportation Fees" if head.is_transport_fee else head.name

                        for i in range(1, inst_count + 1):
                            target_inst = i
                            
                            # Check enrollment (Default is True if no record)
                            is_enrolled = enroll_map.get((student.id, head.id, i), True)
                            if not is_enrolled:
                                continue
                            
                            total_expected += inst_amt
                            
                            # Get paid amount from map
                            paid_amt = trans_map.get((student.id, head.id, i), 0)
                            total_paid += paid_amt
                            
                            if target_inst not in installment_data:
                                installment_data[target_inst] = {'heads': {}}
                                
                            if display_name not in installment_data[target_inst]['heads']:
                                installment_data[target_inst]['heads'][display_name] = {'due': 0, 'paid': 0, 'pending': 0}
                                
                            installment_data[target_inst]['heads'][display_name]['due'] += inst_amt
                            installment_data[target_inst]['heads'][display_name]['paid'] += paid_amt
                            installment_data[target_inst]['heads'][display_name]['pending'] += (inst_amt - paid_amt)
                    except Exception as inner_e:
                        print(f"Skipping head {head.name} for student {student.name}: {str(inner_e)}")
                        continue
                
                balance = total_expected - total_paid
                if show_all or balance > 0.01: # Small float margin
                    fee_detail_list.append({
                        'id': student.id,
                        'student_id': student.student_id,
                        'name': student.name,
                        'student_class': student.student_class,
                        'total_due': round(total_expected, 2),
                        'total_paid': round(total_paid, 2),
                        'pending_amount': round(balance, 2),
                        'installment_data': installment_data
                    })
            
            fee_detail_list.sort(key=lambda x: x['pending_amount'], reverse=True)
            return Response(fee_detail_list)

        except Exception as e:
            import traceback
            print(f"CRITICAL Error in pending_fees: {str(e)}")
            print(traceback.format_exc())
            return Response({'error': str(e), 'traceback': traceback.format_exc()}, status=500)

    @action(detail=True, methods=['get'])
    def ledger(self, request, pk=None):
        student = self.get_object()
        session = request.query_params.get('session')
        
        # Get global settings for current session or latest
        if not session:
            latest_setting = GlobalFeeSetting.objects.all().order_by('-session').first()
            session = latest_setting.session if latest_setting else datetime.now().strftime('%Y-%m-%d')[:4] # fallback

        heads = FeeHead.objects.all()
        if session:
            heads = heads.filter(session=session)
        
        applicable_heads = heads.filter(amounts__class_name=student.student_class)
        
        entries = []
        # Debits: Fee assignments - Group as "All" per head
        for head in applicable_heads:
            if head.is_transport_fee and (not student.has_transport or student.transport_fee_head_id != head.id):
                continue

            try:
                amt = float(FeeAmount.objects.get(fee_head=head, class_name=student.student_class).amount)
                display_name = "Transportation Fees" if head.is_transport_fee else head.name
                entries.append({
                    'date': datetime.now().strftime('%Y-%m-%d'),
                    'description': f"Fee Assigned: {display_name}",
                    'installment': 'All',
                    'debit': amt,
                    'credit': 0,
                })
            except FeeAmount.DoesNotExist:
                continue
        
        # Credits: Payments
        transactions = FeeTransaction.objects.filter(student=student, fee_head__in=applicable_heads).order_by('payment_date')
        for t in transactions:
            display_name = "Transportation Fees" if t.fee_head and t.fee_head.is_transport_fee else (t.fee_head.name if t.fee_head else 'General')
            entries.append({
                'date': t.payment_date.strftime('%Y-%m-%d'),
                'description': f"Payment: {display_name}",
                'installment': t.installment_number,
                'debit': 0,
                'credit': float(t.amount_paid),
            })
        
        # Sort by date
        entries.sort(key=lambda x: x['date'])
        
        # Calculate running sum
        running_sum = 0
        for entry in entries:
            running_sum += (entry['debit'] - entry['credit'])
            entry['balance'] = running_sum

        return Response(entries)

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def bulk_import(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(file_obj)
            sheet = wb.active
            
            # Use list(sheet.iter_rows(min_row=2)) to skip header
            created_count = 0
            updated_count = 0
            errors = []
            
            for row in sheet.iter_rows(min_row=2):
                if not row[0].value: continue
                
                try:
                    s_id = str(row[0].value).strip()
                    name = str(row[1].value).strip()
                    s_class = str(row[2].value).strip()
                    contact = str(row[3].value).strip() if row[3].value else ""
                    has_trans_str = str(row[4].value).strip().lower() if row[4].value else ""
                    trans_head_name = str(row[5].value).strip() if row[5].value else None
                    
                    has_transport = has_trans_str in ['yes', 'true', '1', 'y']
                    
                    trans_head = None
                    if trans_head_name:
                        trans_head = FeeHead.objects.filter(name__iexact=trans_head_name, is_transport_fee=True).first()
                    
                    student, created = Student.objects.update_or_create(
                        student_id=s_id,
                        defaults={
                            'name': name,
                            'student_class': s_class,
                            'contact_number': contact,
                            'has_transport': has_transport,
                            'transport_fee_head': trans_head
                        }
                    )
                    
                    if created: created_count += 1
                    else: updated_count += 1
                except Exception as row_error:
                    errors.append(f"Row {row[0].row}: {str(row_error)}")
                
            return Response({
                'message': f'Import completed. Created: {created_count}, Updated: {updated_count}',
                'created': created_count,
                'updated': updated_count,
                'errors': errors
            })
        except Exception as e:
            return Response({'error': f"Failed to parse Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        
        # Check if status is being changed to TC
        new_status = request.data.get('status')
        if new_status == 'TC' and instance.status != 'TC':
            # Check for dues
            heads = FeeHead.objects.all()
            applicable_heads = heads.filter(amounts__class_name=instance.student_class)
            total_expected = 0
            for head in applicable_heads:
                if head.is_transport_fee and not instance.has_transport: continue
                if head.is_transport_fee and instance.transport_fee_head_id != head.id: continue
                
                try:
                    amt = FeeAmount.objects.get(fee_head=head, class_name=instance.student_class).amount
                    total_expected += amt
                except FeeAmount.DoesNotExist:
                    continue
            
            total_paid = FeeTransaction.objects.filter(student=instance).aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
            balance = total_expected - total_paid
            
            if balance > 0:
                return Response({'error': f'Cannot mark as TC. Student has pending dues: {balance}'}, status=status.HTTP_400_BAD_REQUEST)

        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        kwargs['partial'] = True
        return self.update(request, *args, **kwargs)

    @action(detail=True, methods=['post'])
    def manage_enrollment(self, request, pk=None):
        """
        Manage fee enrollments for a student.
        POST body: {
            "fee_head_id": 1,
            "session": "2026-27",
            "enrollments": {
                "1": true,  // enrolled in installment 1
                "2": true,
                "3": false, // opted out of installment 3
                "4": false
            }
        }
        """
        student = self.get_object()
        fee_head_id = request.data.get('fee_head_id')
        session = request.data.get('session')
        enrollments = request.data.get('enrollments', {})
        
        if not fee_head_id or not session:
            return Response({'error': 'fee_head_id and session are required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            fee_head = FeeHead.objects.get(id=fee_head_id)
        except FeeHead.DoesNotExist:
            return Response({'error': 'Fee head not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Update or create enrollment records
        for inst_num_str, is_enrolled in enrollments.items():
            inst_num = int(inst_num_str)
            
            # Only create record if opted OUT (is_enrolled=False)
            # or if explicitly setting to enrolled (for clarity)
            StudentFeeEnrollment.objects.update_or_create(
                student=student,
                fee_head=fee_head,
                session=session,
                installment_number=inst_num,
                defaults={'is_enrolled': is_enrolled}
            )
        
        return Response({'success': True, 'message': 'Enrollments updated'})
    
    @action(detail=True, methods=['get'])
    def get_enrollments(self, request, pk=None):
        """
        Get all fee enrollments for a student in a session.
        Query params: ?session=2026-27
        """
        student = self.get_object()
        session = request.query_params.get('session')
        
        if not session:
            return Response({'error': 'session parameter is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        enrollments = StudentFeeEnrollment.objects.filter(
            student=student,
            session=session
        ).select_related('fee_head')
        
        from fees.serializers import StudentFeeEnrollmentSerializer
        serializer = StudentFeeEnrollmentSerializer(enrollments, many=True)
        
        return Response(serializer.data)


# Authentication Views
from django.contrib.auth import authenticate, login, logout
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authtoken.models import Token
import json

@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    try:
        data = json.loads(request.body)
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return JsonResponse({'error': 'Username and password are required'}, status=400)
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            token, created = Token.objects.get_or_create(user=user)
            return JsonResponse({
                'success': True,
                'token': token.key,
                'user': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'is_staff': user.is_staff,
                }
            })
        else:
            return JsonResponse({'error': 'Invalid credentials'}, status=401)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def logout_view(request):
    try:
        # Delete the token to log out
        request.user.auth_token.delete()
    except Exception:
        pass
    return JsonResponse({'success': True})

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def check_auth(request):
    return JsonResponse({
        'authenticated': True,
        'user': {
            'id': request.user.id,
            'username': request.user.username,
            'email': request.user.email,
            'is_staff': request.user.is_staff,
        }
    })

